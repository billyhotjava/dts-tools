from __future__ import annotations

import csv
import datetime as dt
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass
from typing import Any, Iterable
from pathlib import Path

from .config import ColumnMapping, ColumnType, ExcelConfig


class ExcelError(RuntimeError):
    pass


@dataclass(frozen=True)
class ParsedRow:
    row_number: int
    values: dict[str, Any]  # db_column -> value
    errors: list[str]

    @property
    def ok(self) -> bool:
        return not self.errors


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _to_bool(value: Any) -> bool | None:
    if _is_empty(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(int(value))
    if isinstance(value, str):
        v = value.strip().lower()
        if v in {"true", "1", "y", "yes", "on", "是", "启用"}:
            return True
        if v in {"false", "0", "n", "no", "off", "否", "禁用"}:
            return False
    raise ValueError(f"Invalid bool: {value!r}")


def convert_value(value: Any, typ: ColumnType) -> Any:
    if _is_empty(value):
        return None

    if typ == "str":
        return str(value).strip()

    if typ == "int":
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if not value.is_integer():
                raise ValueError(f"Expected int, got float={value}")
            return int(value)
        return int(str(value).strip())

    if typ == "float":
        if isinstance(value, bool):
            return float(int(value))
        if isinstance(value, (int, float)):
            return float(value)
        return float(str(value).strip())

    if typ == "decimal":
        if isinstance(value, bool):
            return Decimal(int(value))
        if isinstance(value, int):
            return Decimal(value)
        if isinstance(value, float):
            return Decimal(str(value))
        if isinstance(value, Decimal):
            return value
        try:
            s = str(value).strip()
            s = s.replace(",", "")
            s = s.replace("￥", "").replace("¥", "")
            return Decimal(s)
        except (InvalidOperation, ValueError) as e:
            raise ValueError(f"Invalid decimal: {value!r}") from e

    if typ in {"date", "datetime"}:
        if isinstance(value, dt.datetime):
            return value.date() if typ == "date" else value
        if isinstance(value, dt.date):
            return value if typ == "date" else dt.datetime.combine(value, dt.time.min)
        s = str(value).strip()
        try:
            from dateutil import parser as dt_parser  # type: ignore[import-not-found]

            parsed = dt_parser.parse(s)
            return parsed.date() if typ == "date" else parsed
        except Exception:
            # Fallback for offline/minimal env: try common formats
            fmts = (
                "%Y-%m-%d",
                "%Y/%m/%d",
                "%Y.%m.%d",
                "%Y-%m-%d %H:%M:%S",
                "%Y/%m/%d %H:%M:%S",
                "%Y.%m.%d %H:%M:%S",
            )
            for fmt in fmts:
                try:
                    parsed = dt.datetime.strptime(s, fmt)
                    return parsed.date() if typ == "date" else parsed
                except Exception:
                    pass
            raise ValueError(f"Invalid {typ}: {value!r}")

    if typ == "bool":
        return _to_bool(value)

    raise ValueError(f"Unsupported type: {typ}")


def iter_rows(
    excel_path: str,
    excel_cfg: ExcelConfig,
    columns: list[ColumnMapping],
) -> tuple[list[str], Iterable[ParsedRow]]:
    suffix = Path(excel_path).suffix.lower()
    if suffix == ".csv":
        return _iter_csv_rows(excel_path, excel_cfg, columns)
    return _iter_xlsx_rows(excel_path, excel_cfg, columns)


def _iter_csv_rows(
    file_path: str,
    cfg: ExcelConfig,
    columns: list[ColumnMapping],
) -> tuple[list[str], Iterable[ParsedRow]]:
    header_row = cfg.header_row
    if header_row < 1 or cfg.start_row < 1:
        raise ExcelError("csv header_row/start_row must be >= 1")

    try:
        f = open(file_path, "r", encoding=cfg.csv_encoding, newline="")
    except Exception as e:
        raise ExcelError(f"Failed to read CSV file: {file_path}: {e}") from e

    reader = csv.reader(f, delimiter=cfg.csv_delimiter, quotechar=cfg.csv_quotechar)
    rows = enumerate(reader, start=1)

    header_values: list[str] | None = None
    for row_no, row in rows:
        if row_no == header_row:
            header_values = [str(v).strip() if v is not None else "" for v in row]
            break
    if header_values is None:
        f.close()
        raise ExcelError(f"CSV header_row={header_row} out of range: {file_path}")

    header_index: dict[str, int] = {}
    for idx, name in enumerate(header_values):
        if name and name not in header_index:
            header_index[name] = idx

    missing_headers: list[str] = []
    for c in columns:
        if c.excel not in header_index:
            missing_headers.append(c.excel)
    if missing_headers:
        f.close()
        raise ExcelError(f"Missing CSV headers: {missing_headers}")

    def _gen():
        try:
            for row_no, row in rows:
                if row_no < cfg.start_row:
                    continue
                values: dict[str, Any] = {}
                errors: list[str] = []
                for c in columns:
                    idx = header_index[c.excel]
                    raw = row[idx] if idx < len(row) else None
                    try:
                        converted = convert_value(raw, c.type)
                        if converted is None and c.default is not None:
                            converted = c.default
                        if converted is None and c.required:
                            errors.append(f"{c.excel} required")
                        if isinstance(converted, str) and c.max_length and len(converted) > c.max_length:
                            converted = converted[: c.max_length]
                        values[c.db] = converted
                    except Exception as e:
                        errors.append(f"{c.excel} invalid: {e}")
                        values[c.db] = None
                yield ParsedRow(row_number=row_no, values=values, errors=errors)
        finally:
            f.close()

    return [], _gen()


def _iter_xlsx_rows(
    excel_path: str,
    excel_cfg: ExcelConfig,
    columns: list[ColumnMapping],
) -> tuple[list[str], Iterable[ParsedRow]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception as e:
        raise ExcelError("Reading .xlsx requires `openpyxl` installed") from e

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:  # openpyxl raises varied exceptions
        raise ExcelError(f"Failed to read Excel file: {excel_path}: {e}") from e

    try:
        ws = wb[excel_cfg.sheet] if excel_cfg.sheet else wb.active
    except KeyError as e:
        available = ", ".join(wb.sheetnames)
        raise ExcelError(
            f"Sheet not found: {excel_cfg.sheet!r}; available: {available}"
        ) from e

    header_row = excel_cfg.header_row
    header_cells = ws[header_row]
    header_values = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    header_index: dict[str, int] = {}
    for idx, name in enumerate(header_values):
        if name and name not in header_index:
            header_index[name] = idx

    missing_headers: list[str] = []
    for c in columns:
        if c.excel not in header_index:
            missing_headers.append(c.excel)
    if missing_headers:
        wb.close()
        raise ExcelError(f"Missing Excel headers: {missing_headers}")

    def _gen():
        try:
            for row_number, row_values in enumerate(
                ws.iter_rows(min_row=excel_cfg.start_row, values_only=True),
                start=excel_cfg.start_row,
            ):
                values: dict[str, Any] = {}
                errors: list[str] = []
                for c in columns:
                    idx = header_index[c.excel]
                    raw = row_values[idx] if idx < len(row_values) else None
                    try:
                        converted = convert_value(raw, c.type)
                        if converted is None and c.default is not None:
                            converted = c.default
                        if converted is None and c.required:
                            errors.append(f"{c.excel} required")
                        if isinstance(converted, str) and c.max_length and len(converted) > c.max_length:
                            converted = converted[: c.max_length]
                        values[c.db] = converted
                    except Exception as e:
                        errors.append(f"{c.excel} invalid: {e}")
                        values[c.db] = None

                yield ParsedRow(row_number=row_number, values=values, errors=errors)
        finally:
            wb.close()

    return missing_headers, _gen()
